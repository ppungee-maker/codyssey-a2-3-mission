"""내보내기 — CSV · JSONL 두 포맷.

**두 포맷을 함께 두는 이유**는 소비자가 다르기 때문이다.

  CSV   = 사람·엑셀. 표로 열어 정렬·필터한다. 줄바꿈이 든 값에 약하다.
  JSONL = 프로그램. 한 줄 = 한 레코드라 스트리밍으로 읽고, 중첩·줄바꿈에 강하다.

리뷰 본문에는 줄바꿈이 들어 있을 수 있다. **CSV 에서는 공백으로 펴고, JSONL 에는 원본
그대로** 넣는다.
"""

from __future__ import annotations

import csv
import json
import logging
import os

from . import storage
from .ingest import now_iso

logger = logging.getLogger(__name__)

FIELDS = [
    "review_id", "product", "rating", "created_at", "language",
    "sentiment", "confidence", "text", "cleaned_at", "analyzed_at",
]


def _row_to_dict(row) -> dict:
    """sqlite3.Row → dict. 내보내기 대상 필드만 고른다."""
    return {field: row[field] for field in FIELDS}


def export_csv(rows, out_dir: str) -> str:
    """CSV 로 내보낸다 → 경로.

    `encoding="utf-8-sig"` 를 쓰는 이유: 엑셀(Windows)이 BOM 없는 UTF-8 CSV 를 열면
    한글이 깨진다. BOM 한 글자가 "이건 UTF-8 이다"를 알려 준다.
    """
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"reviews_{now_iso()[:10]}.csv")
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        for row in rows:
            item = _row_to_dict(row)
            # 표 한 칸에 줄바꿈이 들어가면 엑셀에서 행이 밀려 보인다 — 공백으로 편다.
            if item.get("text"):
                item["text"] = str(item["text"]).replace("\n", " ")
            writer.writerow(item)
    logger.info("CSV 저장: %s (%d건)", path, len(rows))
    return path


def export_jsonl(rows, out_dir: str) -> str:
    """JSONL 로 내보낸다 → 경로. 한 줄 = 한 레코드, 원본 줄바꿈 유지."""
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"reviews_{now_iso()[:10]}.jsonl")
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(_row_to_dict(row), ensure_ascii=False) + "\n")
    logger.info("JSONL 저장: %s (%d건)", path, len(rows))
    return path


def export_excel(rows, out_dir: str) -> str:
    """Excel(.xlsx) 로 내보낸다 → 경로.

    openpyxl 을 사용하여 헤더 스타일 및 열 너비를 자동 조정한다.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError(
            "Excel(.xlsx) 로 내보내려면 openpyxl 패키지가 필요합니다. "
            "'pip install openpyxl'을 실행하세요."
        ) from exc

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"reviews_{now_iso()[:10]}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "reviews"

    # 헤더 작성 및 스타일링
    ws.append(FIELDS)
    header_fill = PatternFill(start_color="3D5A80", end_color="3D5A80", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(FIELDS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 데이터 작성
    body_font = Font(name="맑은 고딕", size=10)
    for row in rows:
        item = _row_to_dict(row)
        ws.append([item[f] if item[f] is not None else "" for f in FIELDS])

    # 폰트 적용 및 열 너비 자동 조정
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(FIELDS)):
        for cell in row:
            cell.font = body_font

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            # 한글 등 유니코드 문자는 너비를 1.5배로 계산
            char_len = sum(2 if ord(c) > 127 else 1 for c in val_str[:50])
            max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

    wb.save(path)
    logger.info("Excel 저장: %s (%d건)", path, len(rows))
    return path


def run_export(db_path: str, out_dir: str, fmt: str = "csv", **filters) -> list[str]:
    """내보내기 단계 → 저장된 경로 목록. fmt 는 csv · jsonl · excel · both · all."""
    with storage.connect(db_path) as conn:
        rows = storage.select_clean(conn, **filters)

    if not rows:
        logger.warning("내보낼 데이터가 없습니다 (필터: %s)", filters)
        return []

    paths: list[str] = []
    fmt_lower = fmt.lower()

    if fmt_lower in ("csv", "both", "all"):
        paths.append(export_csv(rows, out_dir))
    if fmt_lower in ("jsonl", "both", "all"):
        paths.append(export_jsonl(rows, out_dir))
    if fmt_lower in ("excel", "xlsx", "all"):
        paths.append(export_excel(rows, out_dir))
    return paths
